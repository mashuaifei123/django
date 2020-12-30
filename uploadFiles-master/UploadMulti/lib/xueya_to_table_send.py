# coding: utf-8
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.styles import Font
import os
import time
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.cell.cell import WriteOnlyCell

G_C_M = ['A', 'B', 'C', 'D', 'E', 'F']  # 列标题名字M性别
G_C_F = ['H', 'I', 'J', 'K', 'L', 'M']  # 列标题名字F性别


def read_csv(path_csv_list):
    '''从一个文件夹路径获取csv文件.拼接返回一个DF, 实验名字, 对照表'''
    csv_list = [fn for fn in os.listdir(path_csv_list) if fn.lower().endswith('.csv')]
    print(csv_list)
    csv_list.sort()
    csv_L_name = list(map(lambda y: path_csv_list + '\\' + y, csv_list))
    fenzu_fname = path_csv_list + '\\' + '分组表.xls'
    df_study = pd.read_csv(open(fenzu_fname), nrows=1)
    study_name = str(df_study.columns[0]).split('for ')[1]
    df_fenzu = pd.read_csv(open(fenzu_fname), skiprows=2)
    df_fenzu.reset_index()
    animal_no = df_fenzu['Study animal number'].tolist()
    animal_id = df_fenzu['Pretest number'].tolist()
    animal_noid = list(zip(animal_no, animal_id))
    df_all = pd.read_csv(open(csv_L_name[0]))
    if len(csv_L_name) > 1:
        for i in range(1, len(csv_L_name)):
            df = pd.read_csv(open(csv_L_name[i]))
            df_all = pd.concat([df_all, df], axis=0)  # 拼接
    df_all = df_all.reset_index()
    print(df_all)
    # df_all.to_csv('sss.csv', encoding='GBK')
    return df_all, animal_noid, study_name



def retuen_jjz(b):
    '''
    返回最接近值所在单元格
    in:[(<Cell 'Sheet'.F7>, 117), (<Cell 'Sheet'.F6>, 119), (<Cell 'Sheet'.F5>,
         122)]
    out:(<Cell 'Sheet'.F7>, <Cell 'Sheet'.F6>)
    '''
    b = [i[0] for i in b]
    a = sorted(b, key=lambda x: x[1])
    if abs(a[0][1]-a[1][1]) < abs(a[1][1]-a[2][1]):
        return a[0][0], a[1][0]
    else:
        return a[1][0], a[2][0]


def id_to_no(id, animal_noid):  # 根据ID返回NO
    for i in animal_noid:
        if id.strip()[1] != 'F' and id.strip()[1] != 'M':
            if id.strip() == str(i[1]):
                return str(i[0])
        if id.strip()[1] == 'F' or id.strip()[1] == 'M':
            return id.strip()
    return '异常ID-{}'.format(id.strip())
    


def no_to_id(id, animal_noid):  # 根据NO返回ID
    for i in animal_noid:
        if id.strip()[1] == 'F' or id.strip()[1] == 'M':
            if id.strip() == str(i[0]):
                return str(i[1])
        if id.strip()[1] != 'F' and id.strip()[1] != 'M':
            return id.strip()
    return '异常ID-{}'.format(id.strip())
    


def df_to_xls_xueya(to_xls_name, Data_Frame, animal_noid, study_name='study_name'):

    error_list = []
    df = Data_Frame

    wb = Workbook()  # 实例化
    ws = wb.active  # 激活一个sheet 默认0 sheet
    Table_Title_M = '{} 雄性动物血压'.format(study_name)
    Table_Title_F = '{} 雌性动物血压'.format(study_name)
    Title_2 = ['动物ID', '动物编号', 'CVDTC', '试验阶段', 'SBP（mmHg）',
               'DBP（mmHg）', 'MBP（mmHg）']

    Title_2_width = 10.50
    Title_2_height = 30.00
    Rows_height = 15.00
    Alig_center = Alignment(horizontal='center', vertical='center',
                            wrapText=True)  # 双剧中
    '''单元格线框颜色设定'''
    border_All = Border(left=Side(border_style='thin', color='FF000000'),
                        right=Side(border_style='thin', color='FF000000'),
                        top=Side(border_style='thin', color='FF000000'),
                        bottom=Side(border_style='thin', color='FF000000'),
                        )
    font_B = Font(name='Times New Roman', size=11, bold=True)

    '''合并单元格,并填充数据Title'''
    ws['A1'].font = Font(name='Times New Roman', size=22, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.cell(row=1, column=1, value=Table_Title_M)

    ws['H1'].font = Font(name='Times New Roman', size=22, bold=True)
    ws['H1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(start_row=1, start_column=9, end_row=1, end_column=15)
    ws.cell(row=1, column=9, value=Table_Title_F)

    '''设定行高列宽'''
    list_Title_2 = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L',
                    'M', 'N', 'O', 'P']
    for i in list_Title_2:
        ws.column_dimensions[i].width = Title_2_width
        ws2_s = ws['{}2'.format(i)]
        ws2_s.font = Font(name='Times New Roman', size=11, bold=True)
        # ws2_s.font = Font(name=u'宋体')
        ws2_s.alignment = Alignment(horizontal='center', vertical='center',
                                    wrapText=True)  # 设定自动换行

    for i in range(1, 3):  # 第1,2行高
        ws.row_dimensions[i].height = Title_2_height

    list_Title_3 = ['C', 'K']
    for i in list_Title_3:
        ws.column_dimensions[i].width = 25.00

    '''填充第2行数据'''
    for i in range(1, 8):
        ws.cell(row=2, column=i, value=Title_2[i-1])
        ws.cell(row=2, column=i,).border = border_All
        ws.cell(row=2, column=i + 8, value=Title_2[i-1])
        ws.cell(row=2, column=i + 8).border = border_All

    '''数据操作填充操作'''
    ''' # 可以使用的第一种情况
    df['试验阶段'] = df[:]['试验阶段'].map(lambda x: str(x).split('-')[1])
    '''
    '''
    df['试验阶段'] = df[:]['试验阶段'].map(lambda x: '{}-{}'.format(str(x).split('-')[1],
                                        str(x).split('-')[-1]))
    '''
    df['ANIMAL_NO'] = df[:]['ANIMAL NO'].map(lambda x: id_to_no(str(x), animal_noid))  # 修改动物编号为 no
    df['Animal_ID'] = df[:]['ANIMAL NO'].map(lambda x: no_to_id(str(x), animal_noid))  # 增加一列动物ID
    df.insert(2, 'M_F', df[:]['ANIMAL_NO'].map(lambda x: str(x).strip()[1:2]))  # 插入新的一列性别
    df['DATE'] = df['DATE'].apply(lambda x : time.strptime('20'+ x, u"%Y/%m/%d"))
    df['DATE'] = df['DATE'].apply(lambda x: time.strftime("%Y-%m-%d", x))
    df['DATE1'] = df.apply(lambda x: x['DATE'] + 'T' + x['TIME'], axis= 1)
    # df.to_csv('df.csv', encoding='GBK')
    df_M = df[df[u'M_F'].isin(['M'])]  # 筛选性别M
    df_M = df_M.sort_values(by=['DATE', 'ANIMAL_NO'])
    # df_M.to_csv('M.csv', encoding='GBK')
    df_F = df[df[u'M_F'].isin(['F'])]  # 筛选性别F
    df_F = df_F.sort_values(by=['DATE', 'ANIMAL_NO'])
    print(df_M)
    # df_F.to_csv('F.csv', encoding='GBK')
    df_gro_M = df_M.groupby(['DATE', 'ANIMAL_NO'])  # 分组
    df_gro_F = df_F.groupby(['DATE', 'ANIMAL_NO'])  # 分组
    start_num_M = 3  # 开始行编号M
    start_num_F = 3  # 开始行编号F

    com_text = '注：mean为标*两次数据的平均值。'
    for k, v in df_gro_M.groups:  # 性别M分组后的组名字
        '''性别M导入excel'''
        cv_l = [(1, 15), (2, 14), (3, 16), (4, 8), (5, 11), (6, 12), (7, 13)]
        df_groups_i = df_gro_M.get_group((k, v))  # 根据组名返回的DataFrame
        # print(df_groups_i)
        df_groups_i_size = df_groups_i.index.size  # DataFrame 行长度
        if df_groups_i_size == 1:  # 只有一条数据情况
            for cvl in cv_l:
                ws.cell(row=start_num_M, column=cvl[0], value=df_groups_i.iloc[0][cvl[1]])
                ws.cell(row=start_num_M, column=cvl[0], ).border = border_All
            start_num_M += 1

        elif df_groups_i_size == 3:  # 只有三条数据情况
            for i in range(0, df_groups_i_size):
                for cvl in cv_l:
                    ws.cell(row=start_num_M, column=cvl[0], value=df_groups_i.iloc[i][cvl[1]])
                    ws.cell(row=start_num_M, column=cvl[0], ).border = border_All
                start_num_M += 1

            '''增加一行为计算平均值切加粗行'''
            ws.cell(row=start_num_M, column=1, value=df_groups_i.iloc[0][15])
            ws.cell(row=start_num_M, column=2, value=df_groups_i.iloc[0][14])
            ws.cell(row=start_num_M, column=3, value=df_groups_i.iloc[0][16])
            ws.cell(row=start_num_M, column=4, value='mean')

            '''根据MBP值近似值返回cell位置数设定平均值公式'''
            dict_Index_col_value = []
            for row in ws.iter_rows(min_row=start_num_M - 3, min_col=7, max_col=7,
                                    max_row=start_num_M - 1):  # 也可以使用openpyxl.worksheet.Worksheet.iter_rows()这个方法
                dict_Index_col_value.append([(cell, cell.value) for cell in row])
            cell_num = retuen_jjz(dict_Index_col_value)
            # print(cell_num)
            cn_1 = int(str(cell_num[0]).split('G')[1][:-1])
            cn_2 = int(str(cell_num[1]).split('G')[1][:-1])
            ws.cell(row=cn_1, column=4, value='{}*'.format(ws['D{}'.format(cn_1)].value))
            ws.cell(row=cn_2, column=4, value='{}*'.format(ws['D{}'.format(cn_2)].value))
            ws['E{}'.format(start_num_M)] = "=AVERAGE(E{},E{})".format(cn_1, cn_2)
            ws.cell(row=start_num_M, column=5).number_format = '0'
            ws['F{}'.format(start_num_M)] = "=AVERAGE(F{},F{})".format(cn_1, cn_2)
            ws.cell(row=start_num_M, column=6).number_format = '0'
            ws['G{}'.format(start_num_M)] = "=AVERAGE(G{},G{})".format(cn_1, cn_2)
            ws.cell(row=start_num_M, column=7).number_format = '0'

            for cvl in cv_l:
                ws.cell(row=start_num_M, column=cvl[0], ).border = border_All
                ws.cell(row=start_num_M, column=cvl[0], ).font = font_B
            start_num_M += 1

        else:
            print('发现异常数据:', k, v)
            error_list.append((k, v))
            pass

    for k, v in df_gro_F.groups:  # 性别F分组后的组名字
        '''性别F导入excel'''
        cv_l = [(9, 15), (10, 14), (11, 16), (12, 8), (13, 11), (14, 12), (15, 13)]
        df_groups_i = df_gro_F.get_group((k, v))  # 根据组名返回的DataFrame
        df_groups_i_size = df_groups_i.index.size  # DataFrame 行长度
        if df_groups_i_size == 1:
            for cvl in cv_l:
                ws.cell(row=start_num_F, column=cvl[0], value=df_groups_i.iloc[0][cvl[1]])
                ws.cell(row=start_num_F, column=cvl[0], ).border = border_All
            start_num_F += 1

        elif df_groups_i_size == 3:
            for i in range(0, df_groups_i_size):
                for cvl in cv_l:
                    ws.cell(row=start_num_F, column=cvl[0], value=df_groups_i.iloc[i][cvl[1]])
                    ws.cell(row=start_num_F, column=cvl[0], ).border = border_All
                start_num_F += 1

            '''增加一行为计算平均值切加粗行'''
            ws.cell(row=start_num_F, column=9, value=df_groups_i.iloc[0][15])
            ws.cell(row=start_num_F, column=10, value=df_groups_i.iloc[0][14])
            ws.cell(row=start_num_F, column=11, value=df_groups_i.iloc[0][16])
            ws.cell(row=start_num_F, column=12, value='mean')

            '''根据MBP值近似值返回cell位置数设定平均值公式'''
            dict_Index_col_value = []
            for row in ws.iter_rows(min_row=start_num_F - 3, min_col=15, max_col=15,
                                    max_row=start_num_F - 1):  # 也可以使用openpyxl.worksheet.Worksheet.iter_rows()这个方法
                dict_Index_col_value.append([(cell, cell.value) for cell in row])
            cell_num = retuen_jjz(dict_Index_col_value)
            cn_1 = int(str(cell_num[0]).split('O')[1][:-1])
            cn_2 = int(str(cell_num[1]).split('O')[1][:-1])
            ws.cell(row=cn_1, column=12, value='{}*'.format(ws['L{}'.format(cn_1)].value))
            ws.cell(row=cn_2, column=12, value='{}*'.format(ws['L{}'.format(cn_2)].value))
            ws['M{}'.format(start_num_F)] = "=AVERAGE(M{},M{})".format(cn_1, cn_2)
            ws.cell(row=start_num_F, column=13).number_format = '0'
            ws['N{}'.format(start_num_F)] = "=AVERAGE(N{},N{})".format(cn_1, cn_2)
            ws.cell(row=start_num_F, column=14).number_format = '0'
            ws['O{}'.format(start_num_F)] = "=AVERAGE(O{},O{})".format(cn_1, cn_2)
            ws.cell(row=start_num_F, column=15).number_format = '0'
            for cvl in cv_l:
                ws.cell(row=start_num_F, column=cvl[0], ).font = font_B
                ws.cell(row=start_num_F, column=cvl[0], ).border = border_All
            start_num_F += 1

        else:
            print('发现异常数据:', k, v)
            error_list.append((k, v))
            pass


    '''最后添加合并行'''
    ws['A{}'.format(start_num_M)].font = Font(name='Times New Roman', size=11)
    ws.cell(row=start_num_M, column=1, value=com_text)
    ws['H{}'.format(start_num_F)].font = Font(name='Times New Roman', size=11)
    ws.cell(row=start_num_F, column=9, value=com_text)
    ws.merge_cells(start_row=start_num_M, start_column=1, end_row=start_num_M, end_column=7)
    ws.merge_cells(start_row=start_num_F, start_column=9, end_row=start_num_F, end_column=15)

    for i in range(1, 8):
        ws.cell(row=start_num_M, column=i, ).border = border_All
    for i in range(9, 15):
        ws.cell(row=start_num_F, column=i, ).border = border_All

    for i in ws.rows:  # 设定所有居中显示
        for k in i:
            k.alignment = Alig_center

    if error_list:  # 判断增加异常数据显示
        ws.cell(row=start_num_M + 3, column=1, value='异常数据:{}'.format(str(error_list)))


    wb.save(to_xls_name)

def xue_ya(csv_path, to_xls_name):
    df_all, animal_noid, study_name = read_csv(csv_path)
    df_to_xls_xueya(to_xls_name + '\\' + '血压.xlsx', df_all, animal_noid, study_name)


if __name__ == "__main__":
    path_csv_list = r'K:\mashuaifei\偏执狂的Django\数据\血压'
    xue_ya(path_csv_list, path_csv_list)
    